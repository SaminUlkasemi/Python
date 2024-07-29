# ---------------------------------------------------------------------------- #
#                              Working with Sheet                              #
# ---------------------------------------------------------------------------- #
# Ref: https://openpyxl.readthedocs.io/en/stable/tutorial.html
import random

import openpyxl # type: ignore
#* Remember that example.xlsx needs to be in the current working directory in 
#* order for you to work with it
import os

print("Current directory: ", os.getcwd())
os.chdir('garbase')
print("Current directory: ", os.getcwd())

'''
wb = openpyxl.load_workbook('example.xlsx')
print("sheetnames: ", wb.sheetnames)

fruitSheet = wb['Fruits']
print("fruitSheet", fruitSheet)
print("fruitSheet type: ", type(fruitSheet))
print("fruitSheet title: ", fruitSheet.title)

print("Active sheet", wb.active)

print(fruitSheet['A1'].value, fruitSheet['B1'].value, fruitSheet['C1'].value)
print("fruitSheet['A1'].row: ", fruitSheet['A1'].row)
print("fruitSheet['A1'].column: ", fruitSheet['A1'].column)
'''

from openpyxl import Workbook

wb = Workbook()
# --------------------------- Get the active sheet --------------------------- #
print("Active sheet: ", wb.active)
# ------------------------- create sheet in Workbook ------------------------- #
item_list   = wb.create_sheet("Item list")          # insert at the end (default)
item_price  = wb.create_sheet("Item price", 0)      # insert at first position
item_rating = wb.create_sheet("Item rating", -1)    # insert at 2nd last

# Chnage the sheet name
item_list.title = 'list'
# get the sheet by namex
print("wb['Item price']: ", wb['Item price'])
# get all the sheet name
print("All sheetnames: ", wb.sheetnames)
# loop through all sheet
for sheet in wb:
    print(sheet)
#? copying worksheet

# ---------------------------------------------------------------------------- #
#                               Playing with data                              #
# ---------------------------------------------------------------------------- #
# set value to the cell
items = ['Apple', 'Mango', 'Orange', 'Banana', 'Pineapple']
for i in range(1, 6):
    item_list[f'A{i}'] = items[i-1]  

for i in range(1, 6):
    item_list[f'B{i}'] = random.randint(5, 100)

for i in range(1, 6):
    item_list[f'C{i}'] = "BDT"


# get the value of a cell
val_b3 = item_list['B3'].value
print(val_b3)

# access using row column
item_list.cell(row=4, column=2, value=1111)

# accessing range of cells
#! find out how to access value using this tuple, what are the methods avaiable
print("item_list['B1':'B5']: ", item_list['B1':'B5']) # return tuple

# access values
print("access values in row: ".center(70, "*"))
for row in item_list.values:
    for val in row:
        print(val)

for row in item_list.iter_rows(max_row=5, max_col=2, values_only=True):
    print(row)


# set value using cell
c = item_list['B1']     # get the cell
c.value = 130     # put the value

print("Accessing row using cell".center(70, "*"))
for i in range(1, 6):
    print(f"{i}: ", item_list.cell(row=i, column=2).value)

# ---------------------------------------------------------------------------- #
#                                save the sheet                                #
# ---------------------------------------------------------------------------- #
#! Warning This operation will overwrite existing files without warning.
wb.save("items.xlsx")
#? ------------------------- save Workbook as template ------------------------ #
wb.template =  True
wb.save('example_template.xltx')


# ---------------------------------------------------------------------------- #
#                             Opening an excel file                            #
# ---------------------------------------------------------------------------- #
print(" Opening a excel file ".center(70, "*"))
from openpyxl import load_workbook

# opening the excel
try:
    wb = load_workbook(filename='items.xlsx')
except Exception as e:
    print(e)

# sheetnames
print("sheetnames: ", wb.sheetnames)

listSheet = wb['list']
print("listSheet: ", listSheet)
print("listSheet type: ", type(listSheet))
print("listSheet title: ", listSheet.title)

print(listSheet['A1'].value, listSheet['B1'].value, listSheet['C1'].value)
print("listSheet['A1'].row: ", listSheet['A1'].row.value)
print("listSheet['A1'].column: ", listSheet['A1'].column)