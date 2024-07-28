# ---------------------------------------------------------------------------- #
#                              Working with Sheet                              #
# ---------------------------------------------------------------------------- #
# Ref: https://openpyxl.readthedocs.io/en/stable/tutorial.html
import random

import openpyxl # type: ignore
#* Remember that example.xlsx needs to be in the current working directory in
#* order for you to work with it
import os

print("Current directory: ", os.getcwd())
os.chdir('garbase')
print("Current directory: ", os.getcwd())

'''
wb = openpyxl.load_workbook('example.xlsx')
print("sheetnames: ", wb.sheetnames)

fruitSheet = wb['Fruits']
print("fruitSheet", fruitSheet)
print("fruitSheet type: ", type(fruitSheet))
print("fruitSheet title: ", fruitSheet.title)

print("Active sheet", wb.active)

print(fruitSheet['A1'].value, fruitSheet['B1'].value, fruitSheet['C1'].value)
print("fruitSheet['A1'].row: ", fruitSheet['A1'].row)
print("fruitSheet['A1'].column: ", fruitSheet['A1'].column)
'''

from openpyxl import Workbook

wb = Workbook()
# --------------------------- Get the active sheet --------------------------- #
print("Active sheet: ", wb.active)
# ------------------------- create sheet in Workbook ------------------------- #
item_list   = wb.create_sheet("Item list")          # insert at the end (default)
item_price  = wb.create_sheet("Item price", 0)      # insert at first position
item_rating = wb.create_sheet("Item rating", -1)    # insert at 2nd last

# Chnage the sheet name
item_list.title = 'list'
# get the sheet by namex
print("wb['Item price']: ", wb['Item price'])
# get all the sheet name
print("All sheetnames: ", wb.sheetnames)
# loop through all sheet
for sheet in wb:
    print(sheet)
#? copying worksheet

# ---------------------------------------------------------------------------- #
#                               Playing with data                              #
# ---------------------------------------------------------------------------- #
# set value to the cell
items = ['Apple', 'Mango', 'Orange', 'Banana', 'Pineapple']
for i in range(1, 6):
    item_list[f'A{i}'] = items[i-1]  

for i in range(1, 6):
    item_list[f'B{i}'] = f"{random.randint(5, 100)} BDT"



# get the value of a cell
val_b3 = item_list['B3'].value
print(val_b3)

# access using row column
item_list.cell(row=4, column=2, value="1111 BDT")

# accessing range of cells
#! find out how to access value using this tuple, what are the methods avaiable
print(item_list['B1':'B5']) # return tuple

# access values
for row in item_list.values:
    for val in row:
        print(val)

for row in item_list.iter_rows(max_row=5, max_col=2, values_only=True):
    print(row)

# ------------------------------ save the sheet ------------------------------ #
wb.save("items.xlsx")